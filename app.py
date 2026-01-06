"""
Analisi Politiche - Applicazione Web
Applicazione Streamlit per l'analisi e il conteggio delle azioni da file Excel.
"""

import streamlit as st
import pandas as pd
from io import BytesIO
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows

# Configurazione pagina
st.set_page_config(
    page_title="Analisi Politiche",
    page_icon="📊",
    layout="wide"
)

# Configurazione di default
DEFAULT_TARIFFE = {
    "A03": 37.14,
    "A06": 35.57,
    "B03": 37.14,
    "B04": 37.14,
    "C06": 499.88
}

DEFAULT_EVENTI_ESCLUSI = ["Annullamento (prima dell'inizio)", "Proposta"]


def load_data(uploaded_file, tariffe: dict, escludi_eventi: list) -> tuple:
    """Carica il file Excel e prepara i dati. Restituisce df valido e df scartato."""
    codici_validi = list(tariffe.keys())
    
    df_raw = pd.read_excel(uploaded_file)
    
    # DataFrame per le righe scartate
    righe_scartate = []
    
    # Crea una copia per lavorare
    df = df_raw.copy()
    
    # Aggiungi colonna per tracciare l'indice originale
    df['_indice_originale'] = df.index + 2  # +2 per header Excel
    
    # 1. Escludi righe con eventi da escludere
    for evento in escludi_eventi:
        mask = df['Evento'] == evento
        scartate = df[mask].copy()
        if len(scartate) > 0:
            scartate['_motivo_esclusione'] = f"Evento escluso: {evento}"
            righe_scartate.append(scartate)
        df = df[~mask]
    
    # Crea una copia per evitare warning pandas
    df = df.copy()
    
    # Estrai codice azione (es. A03, B04, C06)
    df['Codice'] = df['Attività'].str.extract(r'^([A-Z]\d+)')
    
    # 2. Escludi righe con codici non validi (non in tariffe)
    mask_codici_invalidi = ~df['Codice'].isin(codici_validi)
    scartate_codici = df[mask_codici_invalidi].copy()
    if len(scartate_codici) > 0:
        scartate_codici['_motivo_esclusione'] = scartate_codici['Codice'].apply(
            lambda x: f"Codice non in tariffe: {x}" if pd.notna(x) else "Codice non riconosciuto"
        )
        righe_scartate.append(scartate_codici)
    
    df = df[~mask_codici_invalidi].copy()
    
    # Estrai tipo (prima lettera: A, B, C)
    df['Tipo'] = df['Codice'].str[0]
    
    # Determina la data da usare
    df['Data Riferimento'] = df.apply(
        lambda row: row['Data Proposta'] if row['Codice'] == 'C06' else row['Data Fine'],
        axis=1
    )
    
    # Converti in datetime
    df['Data Riferimento'] = pd.to_datetime(df['Data Riferimento'], format='%d/%m/%Y', errors='coerce')
    
    # Estrai anno-mese per aggregazione
    df['Anno-Mese'] = df['Data Riferimento'].dt.to_period('M')
    
    # Combina tutte le righe scartate
    if righe_scartate:
        df_scartate = pd.concat(righe_scartate, ignore_index=True)
    else:
        df_scartate = pd.DataFrame()
    
    return df, df_scartate


def get_operatore_per_persona(df: pd.DataFrame) -> pd.DataFrame:
    """Trova l'operatore più recente per ogni persona."""
    df_sorted = df.sort_values('Data Riferimento', ascending=False)
    operatori = df_sorted.groupby('Destinatario').first()['Operatore'].reset_index()
    operatori.columns = ['Destinatario', 'Operatore']
    return operatori


def conteggio_per_persona_tipo(df: pd.DataFrame, tariffe: dict) -> pd.DataFrame:
    """Conteggio azioni per persona e per tipo."""
    pivot_tipo = pd.pivot_table(
        df, index='Destinatario', columns='Tipo',
        aggfunc='size', fill_value=0
    ).reset_index()
    
    pivot_codice = pd.pivot_table(
        df, index='Destinatario', columns='Codice',
        aggfunc='size', fill_value=0
    ).reset_index()
    
    # Usa i codici dalle tariffe
    codici_dettaglio = list(tariffe.keys())
    codici_presenti = [c for c in codici_dettaglio if c in pivot_codice.columns]
    
    pivot = pivot_tipo.merge(
        pivot_codice[['Destinatario'] + codici_presenti],
        on='Destinatario', how='left'
    )
    
    # Assicura che esistano le colonne tipo (A, B, C) e i codici
    tipi_unici = sorted(df['Tipo'].unique())
    for col in tipi_unici + codici_dettaglio:
        if col not in pivot.columns:
            pivot[col] = 0
    
    operatori = get_operatore_per_persona(df)
    pivot = pivot.merge(operatori, on='Destinatario', how='left')
    
    # Costruisci ordine colonne dinamicamente
    cols_ordinate = ['Destinatario', 'Operatore']
    for tipo in sorted(tipi_unici):
        cols_ordinate.append(tipo)
        # Aggiungi codici di quel tipo
        codici_tipo = [c for c in codici_dettaglio if c.startswith(tipo)]
        cols_ordinate.extend(sorted(codici_tipo))
    
    # Filtra solo colonne esistenti
    cols_ordinate = [c for c in cols_ordinate if c in pivot.columns]
    pivot = pivot[cols_ordinate]
    
    # Calcola totale
    pivot['Totale'] = pivot[[c for c in tipi_unici if c in pivot.columns]].sum(axis=1)
    pivot = pivot.sort_values('Destinatario')
    
    return pivot


def conteggio_per_persona_tipo_mese(df: pd.DataFrame) -> pd.DataFrame:
    """Conteggio azioni per persona, tipo e mese."""
    df_temp = df.copy()
    df_temp['Tipo-Mese'] = df_temp['Tipo'] + '_' + df_temp['Anno-Mese'].astype(str)
    
    pivot = pd.pivot_table(
        df_temp, index='Destinatario', columns='Tipo-Mese',
        aggfunc='size', fill_value=0
    ).reset_index()
    
    operatori = get_operatore_per_persona(df)
    pivot = pivot.merge(operatori, on='Destinatario', how='left')
    
    cols = pivot.columns.tolist()
    cols.remove('Operatore')
    cols.insert(1, 'Operatore')
    pivot = pivot[cols]
    
    return pivot.sort_values('Destinatario')


def conteggio_totale_tipo(df: pd.DataFrame, tariffe: dict) -> pd.DataFrame:
    """Conteggio totale per tipo."""
    conteggio_tipo = df.groupby('Tipo').size().reset_index(name='Conteggio')
    conteggio_codice = df.groupby('Codice').size().reset_index(name='Conteggio')
    conteggio_codice.columns = ['Tipo', 'Conteggio']
    
    risultato = pd.concat([conteggio_tipo, conteggio_codice], ignore_index=True)
    risultato = risultato.sort_values('Tipo')
    
    totale = pd.DataFrame([{'Tipo': 'TOTALE', 'Conteggio': len(df)}])
    risultato = pd.concat([risultato, totale], ignore_index=True)
    
    return risultato


def conteggio_totale_tipo_mese(df: pd.DataFrame) -> pd.DataFrame:
    """Conteggio totale per tipo e mese."""
    pivot = pd.pivot_table(
        df, index='Tipo', columns='Anno-Mese',
        aggfunc='size', fill_value=0
    ).reset_index()
    
    pivot.columns = [str(col) for col in pivot.columns]
    
    mesi_cols = [c for c in pivot.columns if c != 'Tipo']
    totali = pivot[mesi_cols].sum()
    totale_row = pd.DataFrame([['TOTALE'] + totali.tolist()], columns=pivot.columns)
    pivot = pd.concat([pivot, totale_row], ignore_index=True)
    
    return pivot


def conteggio_per_operatore(df: pd.DataFrame, tariffe: dict) -> pd.DataFrame:
    """Conteggio azioni per operatore."""
    pivot = pd.pivot_table(
        df, index='Operatore', columns='Codice',
        aggfunc='size', fill_value=0
    ).reset_index()
    
    codici = list(tariffe.keys())
    for codice in codici:
        if codice not in pivot.columns:
            pivot[codice] = 0
    
    cols = ['Operatore'] + sorted([c for c in codici if c in pivot.columns])
    pivot = pivot[cols]
    pivot['Totale'] = pivot[[c for c in codici if c in pivot.columns]].sum(axis=1)
    
    return pivot.sort_values('Operatore')


def conteggio_per_operatore_mese_tipo(df: pd.DataFrame, tariffe: dict) -> pd.DataFrame:
    """Conteggio azioni per operatore, mese e tipo."""
    df_temp = df.copy()
    df_temp['Mese'] = df_temp['Anno-Mese'].astype(str)
    
    pivot = pd.pivot_table(
        df_temp, index=['Operatore', 'Mese'], columns='Codice',
        aggfunc='size', fill_value=0
    ).reset_index()
    
    codici = list(tariffe.keys())
    for codice in codici:
        if codice not in pivot.columns:
            pivot[codice] = 0
    
    cols = ['Operatore', 'Mese'] + sorted([c for c in codici if c in pivot.columns])
    pivot = pivot[cols]
    pivot['Totale'] = pivot[[c for c in codici if c in pivot.columns]].sum(axis=1)
    
    return pivot.sort_values(['Operatore', 'Mese'])


def calcolo_ricavi(df: pd.DataFrame, tariffe: dict) -> pd.DataFrame:
    """Calcola ricavi per tipologia e mese."""
    df_temp = df.copy()
    df_temp['Mese'] = df_temp['Anno-Mese'].astype(str)
    
    pivot = pd.pivot_table(
        df_temp, index='Mese', columns='Codice',
        aggfunc='size', fill_value=0
    ).reset_index()
    
    for codice in tariffe.keys():
        if codice not in pivot.columns:
            pivot[codice] = 0
    
    cols_codice = sorted(tariffe.keys())
    pivot = pivot[['Mese'] + cols_codice]
    
    for codice, tariffa in tariffe.items():
        pivot[f'{codice}_ricavo'] = pivot[codice] * tariffa
    
    cols_ricavo = [f'{c}_ricavo' for c in cols_codice]
    pivot['Totale_Conteggio'] = pivot[cols_codice].sum(axis=1)
    pivot['Totale_Ricavo'] = pivot[cols_ricavo].sum(axis=1)
    
    return pivot.sort_values('Mese')


def riepilogo_ricavi_per_tipo(df: pd.DataFrame, tariffe: dict) -> pd.DataFrame:
    """Riepilogo ricavi aggregato per tipo."""
    conteggio = df.groupby('Codice').size().reset_index(name='Conteggio')
    
    conteggio['Tariffa (€)'] = conteggio['Codice'].map(tariffe)
    conteggio['Ricavo (€)'] = conteggio['Conteggio'] * conteggio['Tariffa (€)']
    
    totale = pd.DataFrame([{
        'Codice': 'TOTALE',
        'Tariffa (€)': None,
        'Conteggio': conteggio['Conteggio'].sum(),
        'Ricavo (€)': conteggio['Ricavo (€)'].sum()
    }])
    
    risultato = pd.concat([conteggio, totale], ignore_index=True)
    return risultato[['Codice', 'Tariffa (€)', 'Conteggio', 'Ricavo (€)']]


# ============================================================
# FUNZIONI FORMATTAZIONE EXCEL
# ============================================================

def apply_header_style(cell):
    """Applica stile intestazione."""
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    cell.alignment = Alignment(horizontal='center')
    cell.border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )


def apply_cell_style(cell, is_number=False):
    """Applica stile cella."""
    cell.border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    if is_number:
        cell.alignment = Alignment(horizontal='center')


def create_excel_report(df: pd.DataFrame, df_scartate: pd.DataFrame, tariffe: dict) -> bytes:
    """Crea il report Excel con formattazione completa."""
    wb = Workbook()
    
    # ============================================================
    # Foglio 1: Riepilogo
    # ============================================================
    ws_riep = wb.active
    ws_riep.title = "Riepilogo"
    
    row_num = 1
    
    # Titolo
    ws_riep.cell(row=row_num, column=1, value="RIEPILOGO ANALISI POLITICHE")
    ws_riep.cell(row=row_num, column=1).font = Font(bold=True, size=14)
    row_num += 2
    
    # Statistiche generali
    ws_riep.cell(row=row_num, column=1, value="Totale righe analizzate:")
    ws_riep.cell(row=row_num, column=2, value=len(df))
    row_num += 1
    ws_riep.cell(row=row_num, column=1, value="Persone uniche:")
    ws_riep.cell(row=row_num, column=2, value=df['Destinatario'].nunique())
    row_num += 1
    ws_riep.cell(row=row_num, column=1, value="Operatori unici:")
    ws_riep.cell(row=row_num, column=2, value=df['Operatore'].nunique())
    row_num += 1
    ws_riep.cell(row=row_num, column=1, value="Righe scartate:")
    ws_riep.cell(row=row_num, column=2, value=len(df_scartate))
    row_num += 2
    
    # Totali per tipo
    ws_riep.cell(row=row_num, column=1, value="TOTALI PER TIPO")
    ws_riep.cell(row=row_num, column=1).font = Font(bold=True)
    row_num += 1
    df_totale_tipo = conteggio_totale_tipo(df, tariffe)
    for r_idx, row in enumerate(dataframe_to_rows(df_totale_tipo, index=False, header=True), 0):
        for c_idx, value in enumerate(row, 1):
            cell = ws_riep.cell(row=row_num + r_idx, column=c_idx, value=value)
            if r_idx == 0:
                apply_header_style(cell)
            else:
                apply_cell_style(cell, is_number=(c_idx > 1))
                if value == 'TOTALE':
                    cell.font = Font(bold=True)
    row_num += len(df_totale_tipo) + 3
    
    # Totali per tipo e mese
    ws_riep.cell(row=row_num, column=1, value="TOTALI PER TIPO E MESE")
    ws_riep.cell(row=row_num, column=1).font = Font(bold=True)
    row_num += 1
    df_totale_tipo_mese = conteggio_totale_tipo_mese(df)
    for r_idx, row in enumerate(dataframe_to_rows(df_totale_tipo_mese, index=False, header=True), 0):
        for c_idx, value in enumerate(row, 1):
            cell = ws_riep.cell(row=row_num + r_idx, column=c_idx, value=value)
            if r_idx == 0:
                apply_header_style(cell)
            else:
                apply_cell_style(cell, is_number=(c_idx > 1))
                if row[0] == 'TOTALE':
                    cell.font = Font(bold=True)
    row_num += len(df_totale_tipo_mese) + 3
    
    # Riepilogo ricavi
    ws_riep.cell(row=row_num, column=1, value="RIEPILOGO RICAVI")
    ws_riep.cell(row=row_num, column=1).font = Font(bold=True)
    row_num += 1
    df_riepilogo_ricavi = riepilogo_ricavi_per_tipo(df, tariffe)
    for r_idx, row in enumerate(dataframe_to_rows(df_riepilogo_ricavi, index=False, header=True), 0):
        for c_idx, value in enumerate(row, 1):
            cell = ws_riep.cell(row=row_num + r_idx, column=c_idx, value=value)
            if r_idx == 0:
                apply_header_style(cell)
            else:
                apply_cell_style(cell, is_number=(c_idx > 1))
                if c_idx in [2, 4] and isinstance(value, (int, float)):
                    cell.number_format = '#,##0.00 €'
                if row[0] == 'TOTALE':
                    cell.font = Font(bold=True)
    row_num += len(df_riepilogo_ricavi) + 3
    
    # Prime 10 persone
    ws_riep.cell(row=row_num, column=1, value="PRIME 10 PERSONE PER NUMERO AZIONI")
    ws_riep.cell(row=row_num, column=1).font = Font(bold=True)
    row_num += 1
    df_persona = conteggio_per_persona_tipo(df, tariffe).nlargest(10, 'Totale')
    for r_idx, row in enumerate(dataframe_to_rows(df_persona, index=False, header=True), 0):
        for c_idx, value in enumerate(row, 1):
            cell = ws_riep.cell(row=row_num + r_idx, column=c_idx, value=value)
            if r_idx == 0:
                apply_header_style(cell)
            else:
                apply_cell_style(cell, is_number=(c_idx > 2))
    
    ws_riep.column_dimensions['A'].width = 40
    ws_riep.column_dimensions['B'].width = 25
    for col in ['C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O']:
        ws_riep.column_dimensions[col].width = 12
    
    # ============================================================
    # Foglio 2: Per Persona-Tipo
    # ============================================================
    ws1 = wb.create_sheet("Per Persona-Tipo")
    df_persona_tipo = conteggio_per_persona_tipo(df, tariffe)
    
    for r_idx, row in enumerate(dataframe_to_rows(df_persona_tipo, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            cell = ws1.cell(row=r_idx, column=c_idx, value=value)
            if r_idx == 1:
                apply_header_style(cell)
            else:
                apply_cell_style(cell, is_number=(c_idx > 2))
    
    ws1.column_dimensions['A'].width = 40
    ws1.column_dimensions['B'].width = 25
    for col in ['C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O']:
        ws1.column_dimensions[col].width = 10
    
    # ============================================================
    # Foglio 3: Per Persona-Tipo-Mese
    # ============================================================
    ws2 = wb.create_sheet("Per Persona-Tipo-Mese")
    df_persona_tipo_mese = conteggio_per_persona_tipo_mese(df)
    
    for r_idx, row in enumerate(dataframe_to_rows(df_persona_tipo_mese, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            cell = ws2.cell(row=r_idx, column=c_idx, value=value)
            if r_idx == 1:
                apply_header_style(cell)
            else:
                apply_cell_style(cell, is_number=(c_idx > 2))
    
    ws2.column_dimensions['A'].width = 40
    ws2.column_dimensions['B'].width = 25
    
    # ============================================================
    # Foglio 4: Totali per Tipo
    # ============================================================
    ws3 = wb.create_sheet("Totali per Tipo")
    df_totale_tipo = conteggio_totale_tipo(df, tariffe)
    
    for r_idx, row in enumerate(dataframe_to_rows(df_totale_tipo, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            cell = ws3.cell(row=r_idx, column=c_idx, value=value)
            if r_idx == 1:
                apply_header_style(cell)
            else:
                apply_cell_style(cell, is_number=(c_idx > 1))
                if row[0] == 'TOTALE':
                    cell.font = Font(bold=True)
    
    ws3.column_dimensions['A'].width = 15
    ws3.column_dimensions['B'].width = 15
    
    # ============================================================
    # Foglio 5: Totali per Tipo-Mese
    # ============================================================
    ws4 = wb.create_sheet("Totali per Tipo-Mese")
    df_totale_tipo_mese = conteggio_totale_tipo_mese(df)
    
    for r_idx, row in enumerate(dataframe_to_rows(df_totale_tipo_mese, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            cell = ws4.cell(row=r_idx, column=c_idx, value=value)
            if r_idx == 1:
                apply_header_style(cell)
            else:
                apply_cell_style(cell, is_number=(c_idx > 1))
                if row[0] == 'TOTALE':
                    cell.font = Font(bold=True)
    
    ws4.column_dimensions['A'].width = 15
    
    # ============================================================
    # Foglio 6: Per Operatore
    # ============================================================
    ws5 = wb.create_sheet("Per Operatore")
    df_operatore = conteggio_per_operatore(df, tariffe)
    
    for r_idx, row in enumerate(dataframe_to_rows(df_operatore, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            cell = ws5.cell(row=r_idx, column=c_idx, value=value)
            if r_idx == 1:
                apply_header_style(cell)
            else:
                apply_cell_style(cell, is_number=(c_idx > 1))
    
    ws5.column_dimensions['A'].width = 30
    
    # ============================================================
    # Foglio 7: Per Operatore-Mese
    # ============================================================
    ws6 = wb.create_sheet("Per Operatore-Mese")
    df_operatore_mese = conteggio_per_operatore_mese_tipo(df, tariffe)
    
    for r_idx, row in enumerate(dataframe_to_rows(df_operatore_mese, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            cell = ws6.cell(row=r_idx, column=c_idx, value=value)
            if r_idx == 1:
                apply_header_style(cell)
            else:
                apply_cell_style(cell, is_number=(c_idx > 2))
    
    ws6.column_dimensions['A'].width = 30
    ws6.column_dimensions['B'].width = 12
    
    # ============================================================
    # Foglio 8: Ricavi per Mese
    # ============================================================
    ws7 = wb.create_sheet("Ricavi per Mese")
    df_ricavi = calcolo_ricavi(df, tariffe)
    
    for r_idx, row in enumerate(dataframe_to_rows(df_ricavi, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            cell = ws7.cell(row=r_idx, column=c_idx, value=value)
            if r_idx == 1:
                apply_header_style(cell)
            else:
                apply_cell_style(cell, is_number=(c_idx > 1))
                # Formatta colonne ricavo
                col_name = df_ricavi.columns[c_idx - 1]
                if '_ricavo' in str(col_name) or col_name == 'Totale_Ricavo':
                    cell.number_format = '#,##0.00 €'
    
    ws7.column_dimensions['A'].width = 12
    
    # ============================================================
    # Foglio 9: Righe Scartate
    # ============================================================
    ws8 = wb.create_sheet("Righe Scartate")
    
    if len(df_scartate) > 0:
        # Seleziona colonne rilevanti
        cols_to_show = ['_indice_originale', 'Destinatario', 'Operatore', 'Attività', 'Evento', '_motivo_esclusione']
        cols_available = [c for c in cols_to_show if c in df_scartate.columns]
        df_scartate_export = df_scartate[cols_available].copy()
        df_scartate_export.columns = ['Riga Excel', 'Destinatario', 'Operatore', 'Attività', 'Evento', 'Motivo Esclusione']
        
        for r_idx, row in enumerate(dataframe_to_rows(df_scartate_export, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                cell = ws8.cell(row=r_idx, column=c_idx, value=value)
                if r_idx == 1:
                    apply_header_style(cell)
                else:
                    apply_cell_style(cell)
        
        ws8.column_dimensions['A'].width = 12
        ws8.column_dimensions['B'].width = 35
        ws8.column_dimensions['C'].width = 25
        ws8.column_dimensions['D'].width = 30
        ws8.column_dimensions['E'].width = 25
        ws8.column_dimensions['F'].width = 35
    else:
        ws8.cell(row=1, column=1, value="Nessuna riga scartata")
    
    # Salva in BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


# ============================================================
# INTERFACCIA STREAMLIT
# ============================================================

st.title("📊 Analisi Politiche")
st.markdown("Carica un file Excel per analizzare conteggi e ricavi delle azioni.")

# Sidebar per configurazione
with st.sidebar:
    st.header("⚙️ Configurazione")
    
    # --- TARIFFE DINAMICHE ---
    st.subheader("💰 Tariffe (€)")
    
    # Inizializza le tariffe nello stato della sessione
    if 'tariffe' not in st.session_state:
        st.session_state.tariffe = DEFAULT_TARIFFE.copy()
    
    # Form per aggiungere nuova tariffa
    with st.expander("➕ Aggiungi nuova tariffa"):
        col1, col2 = st.columns(2)
        with col1:
            nuovo_codice = st.text_input("Codice (es. A07)", key="nuovo_codice", max_chars=5)
        with col2:
            nuova_tariffa = st.number_input("Tariffa €", value=0.0, step=0.01, key="nuova_tariffa")
        
        if st.button("Aggiungi", key="btn_aggiungi"):
            if nuovo_codice and nuovo_codice.strip():
                codice = nuovo_codice.strip().upper()
                st.session_state.tariffe[codice] = nuova_tariffa
                st.success(f"Aggiunta tariffa {codice}: €{nuova_tariffa:.2f}")
                st.rerun()
    
    # Mostra tariffe esistenti con possibilità di modifica/rimozione
    st.markdown("**Tariffe attuali:**")
    tariffe_da_rimuovere = []
    
    for codice in sorted(st.session_state.tariffe.keys()):
        col1, col2, col3 = st.columns([2, 2, 1])
        with col1:
            st.text(codice)
        with col2:
            nuova_val = st.number_input(
                f"tariffa_{codice}",
                value=st.session_state.tariffe[codice],
                step=0.01,
                format="%.2f",
                key=f"tariffa_{codice}",
                label_visibility="collapsed"
            )
            st.session_state.tariffe[codice] = nuova_val
        with col3:
            if st.button("🗑️", key=f"rimuovi_{codice}"):
                tariffe_da_rimuovere.append(codice)
    
    # Rimuovi tariffe marcate
    for codice in tariffe_da_rimuovere:
        del st.session_state.tariffe[codice]
        st.rerun()
    
    # Pulsante reset tariffe
    if st.button("🔄 Ripristina predefinite"):
        st.session_state.tariffe = DEFAULT_TARIFFE.copy()
        st.rerun()
    
    st.divider()
    
    # --- EVENTI DA ESCLUDERE ---
    st.subheader("🚫 Eventi da Escludere")
    eventi_input = st.text_area(
        "Un evento per riga",
        value="\n".join(DEFAULT_EVENTI_ESCLUSI),
        height=100,
        key="eventi_esclusi"
    )
    escludi_eventi = [e.strip() for e in eventi_input.split("\n") if e.strip()]

# Usa le tariffe dallo stato della sessione
tariffe = st.session_state.tariffe

# Upload file
uploaded_file = st.file_uploader(
    "Carica file Excel (.xls o .xlsx)",
    type=['xls', 'xlsx']
)

if uploaded_file is not None:
    try:
        with st.spinner("Analisi in corso..."):
            df, df_scartate = load_data(uploaded_file, tariffe, escludi_eventi)
        
        # Metriche principali
        st.subheader("📈 Riepilogo")
        col1, col2, col3, col4, col5 = st.columns(5)
        with col1:
            st.metric("Righe Analizzate", len(df))
        with col2:
            st.metric("Persone Uniche", df['Destinatario'].nunique())
        with col3:
            st.metric("Operatori", df['Operatore'].nunique())
        with col4:
            st.metric("Righe Scartate", len(df_scartate))
        with col5:
            ricavi_totali = (df.groupby('Codice').size() * df['Codice'].map(tariffe)).sum()
            st.metric("Ricavi Totali", f"€ {ricavi_totali:,.2f}")
        
        # Tabs per le diverse viste
        tab1, tab2, tab3, tab4, tab5, tab6, tab7 = st.tabs([
            "📋 Per Persona",
            "📅 Per Persona/Mese",
            "📊 Totali",
            "👤 Per Operatore",
            "💰 Ricavi",
            "📈 Ricavi/Mese",
            "🚫 Righe Scartate"
        ])
        
        with tab1:
            st.subheader("Conteggio per Persona e Tipo")
            df_persona = conteggio_per_persona_tipo(df, tariffe)
            st.dataframe(df_persona, use_container_width=True, hide_index=True)
        
        with tab2:
            st.subheader("Conteggio per Persona, Tipo e Mese")
            df_persona_mese = conteggio_per_persona_tipo_mese(df)
            st.dataframe(df_persona_mese, use_container_width=True, hide_index=True)
        
        with tab3:
            col_a, col_b = st.columns(2)
            with col_a:
                st.subheader("Totali per Tipo")
                st.dataframe(conteggio_totale_tipo(df, tariffe), use_container_width=True, hide_index=True)
            with col_b:
                st.subheader("Totali per Tipo e Mese")
                st.dataframe(conteggio_totale_tipo_mese(df), use_container_width=True, hide_index=True)
        
        with tab4:
            col_a, col_b = st.columns(2)
            with col_a:
                st.subheader("Per Operatore")
                st.dataframe(conteggio_per_operatore(df, tariffe), use_container_width=True, hide_index=True)
            with col_b:
                st.subheader("Per Operatore e Mese")
                st.dataframe(conteggio_per_operatore_mese_tipo(df, tariffe), use_container_width=True, hide_index=True)
        
        with tab5:
            st.subheader("Riepilogo Ricavi per Tipo")
            st.dataframe(riepilogo_ricavi_per_tipo(df, tariffe), use_container_width=True, hide_index=True)
        
        with tab6:
            st.subheader("Ricavi per Mese")
            st.dataframe(calcolo_ricavi(df, tariffe), use_container_width=True, hide_index=True)
        
        with tab7:
            st.subheader("Righe Scartate")
            if len(df_scartate) > 0:
                # Raggruppa per motivo
                motivi = df_scartate['_motivo_esclusione'].value_counts()
                st.markdown("**Riepilogo motivi di esclusione:**")
                for motivo, count in motivi.items():
                    st.write(f"- {motivo}: **{count}** righe")
                
                st.divider()
                
                # Mostra dettaglio
                cols_to_show = ['_indice_originale', 'Destinatario', 'Operatore', 'Attività', 'Evento', '_motivo_esclusione']
                cols_available = [c for c in cols_to_show if c in df_scartate.columns]
                df_scartate_display = df_scartate[cols_available].copy()
                df_scartate_display.columns = ['Riga Excel', 'Destinatario', 'Operatore', 'Attività', 'Evento', 'Motivo Esclusione']
                
                st.dataframe(df_scartate_display, use_container_width=True, hide_index=True)
            else:
                st.success("✅ Nessuna riga scartata!")
        
        # Download Excel
        st.markdown("---")
        
        # Prepara Excel con formattazione
        excel_data = create_excel_report(df, df_scartate, tariffe)
        
        timestamp = datetime.now().strftime("%Y%m%d_%H%M")
        filename = f"{timestamp}_export_analisi.xlsx"
        
        st.download_button(
            label="📥 Scarica Report Excel",
            data=excel_data,
            file_name=filename,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        
    except Exception as e:
        st.error(f"Errore durante l'analisi: {str(e)}")
        st.exception(e)

else:
    st.info("👆 Carica un file Excel per iniziare l'analisi")
    
    with st.expander("ℹ️ Informazioni sul formato del file"):
        st.markdown("""
        Il file Excel deve contenere le seguenti colonne:
        - **Destinatario**: Nome della persona
        - **Operatore**: Nome dell'operatore
        - **Attività**: Codice attività (es. A03, B04, C06)
        - **Evento**: Stato dell'attività
        - **Data Fine**: Data di completamento
        - **Data Proposta**: Data proposta (usata per C06)
        """)
    
    with st.expander("📋 Tariffe configurate"):
        st.markdown("Tariffe attualmente configurate (modificabili nella barra laterale):")
        for codice, tariffa in sorted(st.session_state.tariffe.items()):
            st.write(f"- **{codice}**: €{tariffa:.2f}")