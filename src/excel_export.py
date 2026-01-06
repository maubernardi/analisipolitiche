"""
Export dei dati in formato Excel con formattazione professionale e grafici.
"""

import pandas as pd
from io import BytesIO
from typing import Dict, List
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from openpyxl.chart import LineChart, BarChart, PieChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.series import DataPoint

from .analysis import AnalisiPolitiche


class ExcelExporter:
    """Esporta i dati in formato Excel con formattazione e grafici."""
    
    # Stili
    HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    HEADER_FONT = Font(bold=True, color="FFFFFF")
    BORDER = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Colori per i grafici
    CHART_COLORS = ['4472C4', 'ED7D31', '70AD47', 'FFC000', '5B9BD5', 'A5A5A5']
    
    def __init__(self, analisi: AnalisiPolitiche, df_scartate: pd.DataFrame):
        """
        Inizializza l'exporter.
        
        Args:
            analisi: Istanza di AnalisiPolitiche con i dati
            df_scartate: DataFrame con le righe scartate
        """
        self.analisi = analisi
        self.df_scartate = df_scartate
    
    def _apply_header_style(self, cell) -> None:
        """Applica stile intestazione."""
        cell.font = self.HEADER_FONT
        cell.fill = self.HEADER_FILL
        cell.alignment = Alignment(horizontal='center')
        cell.border = self.BORDER
    
    def _apply_cell_style(self, cell, is_number: bool = False) -> None:
        """Applica stile cella."""
        cell.border = self.BORDER
        if is_number:
            cell.alignment = Alignment(horizontal='center')
    
    def _write_dataframe(self, ws, df: pd.DataFrame, start_row: int = 1, 
                         num_cols_start: int = 2, bold_total_row: bool = False) -> int:
        """
        Scrive un DataFrame in un foglio Excel con formattazione.
        
        Returns:
            Numero di righe scritte
        """
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 0):
            for c_idx, value in enumerate(row, 1):
                cell = ws.cell(row=start_row + r_idx, column=c_idx, value=value)
                if r_idx == 0:
                    self._apply_header_style(cell)
                else:
                    self._apply_cell_style(cell, is_number=(c_idx >= num_cols_start))
                    if bold_total_row and row[0] == 'TOTALE':
                        cell.font = Font(bold=True)
        
        return len(df) + 1  # +1 per header
    
    def _create_line_chart(self, ws_data, data_start_row: int, data_end_row: int, 
                           num_series: int, title: str) -> LineChart:
        """
        Crea un grafico a linee per l'andamento mensile.
        
        Args:
            ws_data: Worksheet con i dati
            data_start_row: Riga iniziale dei dati
            data_end_row: Riga finale dei dati
            num_series: Numero di serie dati (tipi + totale)
            title: Titolo del grafico
        """
        chart = LineChart()
        chart.title = title
        chart.style = 10
        chart.y_axis.title = "Quantità"
        chart.x_axis.title = "Mese"
        chart.width = 18
        chart.height = 10
        
        # Dati per le serie (colonne dalla 2 in poi)
        data = Reference(ws_data, min_col=2, min_row=data_start_row, 
                        max_col=1 + num_series, max_row=data_end_row)
        
        # Categorie (mesi - colonna 1)
        cats = Reference(ws_data, min_col=1, min_row=data_start_row + 1, 
                        max_row=data_end_row)
        
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        
        # Stile linee
        for i, series in enumerate(chart.series):
            series.smooth = False
            series.marker.symbol = "circle"
            series.marker.size = 7
        
        return chart
    
    def _create_bar_chart(self, ws_data, data_start_row: int, data_end_row: int,
                          title: str) -> BarChart:
        """
        Crea un grafico a barre orizzontali per i ricavi.
        """
        chart = BarChart()
        chart.type = "bar"  # Barre orizzontali
        chart.style = 10
        chart.title = title
        chart.y_axis.title = "Codice Azione"
        chart.x_axis.title = "Ricavo (€)"
        chart.width = 18
        chart.height = 10
        
        # Dati ricavi (colonna 2)
        data = Reference(ws_data, min_col=2, min_row=data_start_row, 
                        max_row=data_end_row)
        
        # Categorie (codici - colonna 1)
        cats = Reference(ws_data, min_col=1, min_row=data_start_row + 1, 
                        max_row=data_end_row)
        
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        
        # Mostra etichette dati
        chart.dataLabels = DataLabelList()
        chart.dataLabels.showVal = True
        chart.dataLabels.numFmt = '€#,##0'
        
        return chart
    
    def _create_pie_chart(self, ws_data, data_start_row: int, data_end_row: int,
                          title: str) -> PieChart:
        """
        Crea un grafico a torta per utenti per operatore.
        """
        chart = PieChart()
        chart.title = title
        chart.width = 14
        chart.height = 10
        
        # Dati (colonna 2)
        data = Reference(ws_data, min_col=2, min_row=data_start_row, 
                        max_row=data_end_row)
        
        # Categorie (operatori - colonna 1)
        cats = Reference(ws_data, min_col=1, min_row=data_start_row + 1, 
                        max_row=data_end_row)
        
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        
        # Mostra percentuali e valori
        chart.dataLabels = DataLabelList()
        chart.dataLabels.showPercent = True
        chart.dataLabels.showVal = True
        chart.dataLabels.showCatName = False
        
        return chart
    
    def export(self) -> bytes:
        """
        Crea il report Excel completo con grafici.
        
        Returns:
            Contenuto del file Excel come bytes
        """
        wb = Workbook()
        
        # ============================================================
        # Foglio 1: Riepilogo
        # ============================================================
        ws_riep = wb.active
        ws_riep.title = "Riepilogo"
        
        row_num = 1
        
        # Titolo
        ws_riep.cell(row=row_num, column=1, value="RIEPILOGO ANALISI POLITICHE")
        ws_riep.cell(row=row_num, column=1).font = Font(bold=True, size=14)
        row_num += 2
        
        # Statistiche generali
        stats = [
            ("Totale righe analizzate:", len(self.analisi.df)),
            ("Persone uniche:", self.analisi.df['Destinatario'].nunique()),
            ("Operatori unici:", self.analisi.df['Operatore'].nunique()),
            ("Ricavi totali:", self.analisi.ricavi_totali()),
            ("Righe scartate:", len(self.df_scartate))
        ]
        for label, value in stats:
            ws_riep.cell(row=row_num, column=1, value=label)
            cell = ws_riep.cell(row=row_num, column=2, value=value)
            if "Ricavi" in label:
                cell.number_format = '#,##0.00 €'
            row_num += 1
        row_num += 1
        
        # Totali per tipo
        ws_riep.cell(row=row_num, column=1, value="TOTALI PER TIPO")
        ws_riep.cell(row=row_num, column=1).font = Font(bold=True)
        row_num += 1
        df_totale_tipo = self.analisi.conteggio_totale_tipo()
        row_num += self._write_dataframe(ws_riep, df_totale_tipo, row_num, 2, True)
        row_num += 2
        
        # Totali per tipo e mese
        ws_riep.cell(row=row_num, column=1, value="TOTALI PER TIPO E MESE")
        ws_riep.cell(row=row_num, column=1).font = Font(bold=True)
        row_num += 1
        df_totale_tipo_mese = self.analisi.conteggio_totale_tipo_mese()
        row_num += self._write_dataframe(ws_riep, df_totale_tipo_mese, row_num, 2, True)
        row_num += 2
        
        # Riepilogo ricavi
        ws_riep.cell(row=row_num, column=1, value="RIEPILOGO RICAVI")
        ws_riep.cell(row=row_num, column=1).font = Font(bold=True)
        row_num += 1
        df_riepilogo_ricavi = self.analisi.riepilogo_ricavi()
        for r_idx, row in enumerate(dataframe_to_rows(df_riepilogo_ricavi, index=False, header=True), 0):
            for c_idx, value in enumerate(row, 1):
                cell = ws_riep.cell(row=row_num + r_idx, column=c_idx, value=value)
                if r_idx == 0:
                    self._apply_header_style(cell)
                else:
                    self._apply_cell_style(cell, is_number=(c_idx > 1))
                    if c_idx in [2, 4] and isinstance(value, (int, float)):
                        cell.number_format = '#,##0.00 €'
                    if row[0] == 'TOTALE':
                        cell.font = Font(bold=True)
        row_num += len(df_riepilogo_ricavi) + 3
        
        # Utenti per operatore
        ws_riep.cell(row=row_num, column=1, value="UTENTI PER OPERATORE")
        ws_riep.cell(row=row_num, column=1).font = Font(bold=True)
        row_num += 1
        df_utenti_op = self.analisi.utenti_per_operatore()
        row_num += self._write_dataframe(ws_riep, df_utenti_op, row_num, 2, True)
        row_num += 2
        
        # Prime 10 persone
        ws_riep.cell(row=row_num, column=1, value="PRIME 10 PERSONE PER NUMERO AZIONI")
        ws_riep.cell(row=row_num, column=1).font = Font(bold=True)
        row_num += 1
        df_top = self.analisi.top_persone(10)
        self._write_dataframe(ws_riep, df_top, row_num, 3)
        
        # Imposta larghezze colonne
        ws_riep.column_dimensions['A'].width = 40
        ws_riep.column_dimensions['B'].width = 25
        for col in ['C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O']:
            ws_riep.column_dimensions[col].width = 12
        
        # ============================================================
        # Foglio Grafici
        # ============================================================
        ws_grafici = wb.create_sheet("Grafici")
        
        # --- Dati per grafico andamento mensile ---
        ws_grafici.cell(row=1, column=1, value="DATI ANDAMENTO MENSILE")
        ws_grafici.cell(row=1, column=1).font = Font(bold=True)
        
        df_andamento = self.analisi.andamento_mensile()
        andamento_start_row = 2
        for r_idx, row in enumerate(dataframe_to_rows(df_andamento, index=False, header=True), 0):
            for c_idx, value in enumerate(row, 1):
                cell = ws_grafici.cell(row=andamento_start_row + r_idx, column=c_idx, value=value)
                if r_idx == 0:
                    self._apply_header_style(cell)
                else:
                    self._apply_cell_style(cell, is_number=(c_idx > 1))
        
        andamento_end_row = andamento_start_row + len(df_andamento)
        num_series = len(df_andamento.columns) - 1  # Escludi colonna Mese
        
        # Crea grafico a linee
        chart_line = self._create_line_chart(
            ws_grafici, andamento_start_row, andamento_end_row,
            num_series, "Andamento Mensile Azioni"
        )
        ws_grafici.add_chart(chart_line, "H2")
        
        # --- Dati per grafico ricavi ---
        ricavi_start_row = andamento_end_row + 4
        ws_grafici.cell(row=ricavi_start_row - 1, column=1, value="DATI RICAVI PER CODICE")
        ws_grafici.cell(row=ricavi_start_row - 1, column=1).font = Font(bold=True)
        
        df_ricavi = self.analisi.ricavi_per_codice()
        # Prepara dati semplificati per il grafico
        ws_grafici.cell(row=ricavi_start_row, column=1, value="Codice")
        ws_grafici.cell(row=ricavi_start_row, column=2, value="Ricavo (€)")
        self._apply_header_style(ws_grafici.cell(row=ricavi_start_row, column=1))
        self._apply_header_style(ws_grafici.cell(row=ricavi_start_row, column=2))
        
        for i, (_, row) in enumerate(df_ricavi.iterrows(), 1):
            ws_grafici.cell(row=ricavi_start_row + i, column=1, value=row['Codice'])
            cell = ws_grafici.cell(row=ricavi_start_row + i, column=2, value=row['Ricavo'])
            cell.number_format = '#,##0.00 €'
            self._apply_cell_style(ws_grafici.cell(row=ricavi_start_row + i, column=1))
            self._apply_cell_style(ws_grafici.cell(row=ricavi_start_row + i, column=2), True)
        
        ricavi_end_row = ricavi_start_row + len(df_ricavi)
        
        # Crea grafico a barre orizzontali
        chart_bar = self._create_bar_chart(
            ws_grafici, ricavi_start_row, ricavi_end_row,
            "Ricavi per Tipo Azione"
        )
        ws_grafici.add_chart(chart_bar, "H" + str(ricavi_start_row - 1))
        
        # --- Dati per grafico utenti per operatore ---
        utenti_start_row = ricavi_end_row + 4
        ws_grafici.cell(row=utenti_start_row - 1, column=1, value="DATI UTENTI PER OPERATORE")
        ws_grafici.cell(row=utenti_start_row - 1, column=1).font = Font(bold=True)
        
        df_utenti = self.analisi.utenti_per_operatore()
        # Escludi riga TOTALE per il grafico a torta
        df_utenti_grafico = df_utenti[df_utenti['Operatore'] != 'TOTALE']
        
        ws_grafici.cell(row=utenti_start_row, column=1, value="Operatore")
        ws_grafici.cell(row=utenti_start_row, column=2, value="Numero Utenti")
        self._apply_header_style(ws_grafici.cell(row=utenti_start_row, column=1))
        self._apply_header_style(ws_grafici.cell(row=utenti_start_row, column=2))
        
        for i, (_, row) in enumerate(df_utenti_grafico.iterrows(), 1):
            ws_grafici.cell(row=utenti_start_row + i, column=1, value=row['Operatore'])
            ws_grafici.cell(row=utenti_start_row + i, column=2, value=row['Numero Utenti'])
            self._apply_cell_style(ws_grafici.cell(row=utenti_start_row + i, column=1))
            self._apply_cell_style(ws_grafici.cell(row=utenti_start_row + i, column=2), True)
        
        utenti_end_row = utenti_start_row + len(df_utenti_grafico)
        
        # Crea grafico a torta
        chart_pie = self._create_pie_chart(
            ws_grafici, utenti_start_row, utenti_end_row,
            "Utenti per Operatore"
        )
        ws_grafici.add_chart(chart_pie, "H" + str(utenti_start_row - 1))
        
        # Imposta larghezze colonne foglio grafici
        ws_grafici.column_dimensions['A'].width = 25
        ws_grafici.column_dimensions['B'].width = 15
        for col in ['C', 'D', 'E', 'F']:
            ws_grafici.column_dimensions[col].width = 12
        
        # ============================================================
        # Foglio 2: Per Persona-Tipo
        # ============================================================
        ws1 = wb.create_sheet("Per Persona-Tipo")
        df_persona_tipo = self.analisi.conteggio_per_persona_tipo()
        self._write_dataframe(ws1, df_persona_tipo, 1, 3)
        ws1.column_dimensions['A'].width = 40
        ws1.column_dimensions['B'].width = 25
        for col in ['C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O']:
            ws1.column_dimensions[col].width = 10
        
        # ============================================================
        # Foglio 3: Per Persona-Tipo-Mese
        # ============================================================
        ws2 = wb.create_sheet("Per Persona-Tipo-Mese")
        df_persona_tipo_mese = self.analisi.conteggio_per_persona_tipo_mese()
        self._write_dataframe(ws2, df_persona_tipo_mese, 1, 3)
        ws2.column_dimensions['A'].width = 40
        ws2.column_dimensions['B'].width = 25
        
        # ============================================================
        # Foglio 4: Totali per Tipo
        # ============================================================
        ws3 = wb.create_sheet("Totali per Tipo")
        df_totale_tipo = self.analisi.conteggio_totale_tipo()
        self._write_dataframe(ws3, df_totale_tipo, 1, 2, True)
        ws3.column_dimensions['A'].width = 15
        ws3.column_dimensions['B'].width = 15
        
        # ============================================================
        # Foglio 5: Totali per Tipo-Mese
        # ============================================================
        ws4 = wb.create_sheet("Totali per Tipo-Mese")
        df_totale_tipo_mese = self.analisi.conteggio_totale_tipo_mese()
        self._write_dataframe(ws4, df_totale_tipo_mese, 1, 2, True)
        ws4.column_dimensions['A'].width = 15
        
        # ============================================================
        # Foglio 6: Per Operatore
        # ============================================================
        ws5 = wb.create_sheet("Per Operatore")
        df_operatore = self.analisi.conteggio_per_operatore()
        self._write_dataframe(ws5, df_operatore, 1, 2)
        ws5.column_dimensions['A'].width = 30
        
        # ============================================================
        # Foglio 7: Per Operatore-Mese
        # ============================================================
        ws6 = wb.create_sheet("Per Operatore-Mese")
        df_operatore_mese = self.analisi.conteggio_per_operatore_mese()
        self._write_dataframe(ws6, df_operatore_mese, 1, 3)
        ws6.column_dimensions['A'].width = 30
        ws6.column_dimensions['B'].width = 12
        
        # ============================================================
        # Foglio 8: Ricavi per Mese
        # ============================================================
        ws7 = wb.create_sheet("Ricavi per Mese")
        df_ricavi_mese = self.analisi.calcolo_ricavi_per_mese()
        for r_idx, row in enumerate(dataframe_to_rows(df_ricavi_mese, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                cell = ws7.cell(row=r_idx, column=c_idx, value=value)
                if r_idx == 1:
                    self._apply_header_style(cell)
                else:
                    self._apply_cell_style(cell, is_number=(c_idx > 1))
                    col_name = df_ricavi_mese.columns[c_idx - 1]
                    if '_ricavo' in str(col_name) or col_name == 'Totale_Ricavo':
                        cell.number_format = '#,##0.00 €'
        ws7.column_dimensions['A'].width = 12
        
        # ============================================================
        # Foglio 9: Righe Scartate
        # ============================================================
        ws8 = wb.create_sheet("Righe Scartate")
        
        if len(self.df_scartate) > 0:
            cols_to_show = [
                '_indice_originale', 'Destinatario', 'Operatore', 
                'Attività', 'Evento', '_motivo_esclusione'
            ]
            cols_available = [c for c in cols_to_show if c in self.df_scartate.columns]
            df_scartate_export = self.df_scartate[cols_available].copy()
            df_scartate_export.columns = [
                'Riga Excel', 'Destinatario', 'Operatore', 
                'Attività', 'Evento', 'Motivo Esclusione'
            ][:len(cols_available)]
            
            self._write_dataframe(ws8, df_scartate_export, 1, 7)
            
            ws8.column_dimensions['A'].width = 12
            ws8.column_dimensions['B'].width = 35
            ws8.column_dimensions['C'].width = 25
            ws8.column_dimensions['D'].width = 30
            ws8.column_dimensions['E'].width = 25
            ws8.column_dimensions['F'].width = 35
        else:
            ws8.cell(row=1, column=1, value="Nessuna riga scartata")
        
        # Salva in BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()
